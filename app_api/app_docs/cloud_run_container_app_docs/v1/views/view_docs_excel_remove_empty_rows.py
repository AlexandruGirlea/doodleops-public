import os
import logging

import aiofiles
from openpyxl import load_workbook
from fastapi import APIRouter
from fastapi import (
    File, UploadFile, HTTPException, BackgroundTasks, Depends
)
from fastapi.responses import FileResponse

from views.urls import urls
from access_management.api_auth import verify_token
from utils.helper_methods import (
    cleanup_temp_dir, get_temp_file_path, validate_doc_file_input,
    EXCEL_MEDIA_TYPE
)


logger = logging.getLogger(__name__)

docs_excel_remove_empty_rows_router = APIRouter(
    tags=["Excel"],
    responses={404: {"description": "Not found"}},
)


@docs_excel_remove_empty_rows_router.post(
    urls.get("excel").get("remove_empty_rows"),
    include_in_schema=True,
)
async def remove_empty_rows(
        background_tasks: BackgroundTasks,
        file: UploadFile = File(...),
        token_data: bool = Depends(verify_token),
):
    """
    Removes rows that are completely empty from an Excel file.
    """
    validate_doc_file_input(file=file, file_extensions=('xlsx', 'xls'))

    input_file_path = get_temp_file_path(extension='xlsx')
    temp_dir = os.path.dirname(input_file_path)

    try:
        async with aiofiles.open(input_file_path, 'wb') as input_file:
            content = await file.read()
            await input_file.write(content)

        try:
            workbook = load_workbook(filename=input_file_path, data_only=False)
        except:
            raise HTTPException(
                status_code=400,
                detail="Could not read the file"
            )

        for sheet_name in workbook.sheetnames:
            # Get the sheet
            sheet = workbook[sheet_name]

            # Collect indices of rows to delete
            rows_to_delete = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                       min_col=1, max_col=sheet.max_column):
                if all(cell.value is None for cell in row):
                    rows_to_delete.append(row[0].row)

            # Delete rows in reverse order
            for row_index in reversed(rows_to_delete):
                sheet.delete_rows(row_index, 1)

        output_file_path = os.path.join(temp_dir, 'output.xlsx')

        workbook.save(output_file_path)
        background_tasks.add_task(cleanup_temp_dir, temp_dir=temp_dir)
        return FileResponse(
            output_file_path,
            media_type=EXCEL_MEDIA_TYPE,
            filename='output.xlsx'
        )

    except (OSError, HTTPException, Exception) as e:
        logging.error(f"Error: {e}")
        cleanup_temp_dir(file_path=input_file_path)

        if isinstance(e, HTTPException):
            raise HTTPException(
                status_code=e.status_code,
                detail="Could not process the file"
            )

        raise HTTPException(
            status_code=500,
            detail="Server error"
        )
