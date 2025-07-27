import os
import logging

import aiofiles
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from fastapi import APIRouter
from fastapi import File, UploadFile, HTTPException, BackgroundTasks, Depends
from fastapi.responses import FileResponse

from views.urls import urls
from access_management.api_auth import verify_token
from utils.helper_methods import (
	cleanup_temp_dir, get_temp_file_path, validate_doc_file_input,
	archive_to_zip
)

logger = logging.getLogger(__name__)

docs_excel_extract_each_sheet_to_new_excels_router = APIRouter(
	tags=["Excel"],
	responses={404: {"description": "Not found"}},
)


@docs_excel_extract_each_sheet_to_new_excels_router.post(
	urls.get("excel").get("extract_each_sheet_to_new_excels"),
	include_in_schema=True,
)
async def extract_each_sheet_to_new_excels(
		background_tasks: BackgroundTasks,
		file: UploadFile = File(...),
		token_data: bool = Depends(verify_token),
):
	"""
	Attention: This API does not keep the cell formatting. It only copies the cell
	values to new Excel files.
	"""
	validate_doc_file_input(file=file, file_extensions=('xlsx', 'xls'))
	input_file_path = get_temp_file_path(extension='xlsx')
	temp_dir = os.path.dirname(input_file_path)

	async with aiofiles.open(input_file_path, 'wb') as input_file:
		content = await file.read()
		await input_file.write(content)

	try:
		workbook = load_workbook(filename=input_file_path, read_only=True)
	except Exception as e:
		logger.error(f"Could not read the file: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise HTTPException(
			status_code=400,
			detail="Could not read the file."
		)

	try:
		if len(workbook.sheetnames) == 1:
			raise HTTPException(
				status_code=400,
				detail=(
					f"The file has only one sheet `{workbook.sheetnames[0]}`. "
					"Please upload a file with multiple sheets."
				)
			)

		for sheet_name in workbook.sheetnames:
			# Get the sheet
			sheet = workbook[sheet_name]

			# Create a new workbook and add the sheet
			new_workbook = Workbook()
			new_sheet = new_workbook.active
			new_sheet.title = sheet_name

			# Copy the content and formatting from the original sheet to the new sheet
			for row in sheet.iter_rows():
				for cell in row:
					if cell.value is None:
						continue  # Skip empty cells
					new_sheet.cell(
						row=cell.row, column=cell.column, value=cell.value
					)
			# Save the new workbook
			output_file_path = os.path.join(temp_dir, f"sheet_{sheet_name}.xlsx")
			new_workbook.save(output_file_path)

		zip_file = archive_to_zip(
			temp_dir_path=temp_dir,
			zip_file_name="sheets.zip",
			file_starts_with="sheet_"
		)

		background_tasks.add_task(cleanup_temp_dir, temp_dir=temp_dir)

		return FileResponse(
			zip_file, filename="sheets.zip", media_type="application/zip"
		)
	except HTTPException as e:
		logger.error(f"Error: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise e
	except Exception as e:
		logger.error(f"Error: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise HTTPException(status_code=500, detail="Server error")
