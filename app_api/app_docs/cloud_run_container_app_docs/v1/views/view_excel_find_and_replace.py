import os
import logging

import aiofiles
from openpyxl import load_workbook
from fastapi import APIRouter
from fastapi import (
	File, UploadFile, HTTPException, BackgroundTasks, Depends, Form
)
from fastapi.responses import FileResponse

from views.urls import urls
from access_management.api_auth import verify_token
from utils.helper_methods import (
	cleanup_temp_dir, get_temp_file_path, validate_doc_file_input,
	EXCEL_MEDIA_TYPE,
)

logger = logging.getLogger(__name__)

docs_find_and_replace_router = APIRouter(
	tags=["Excel"],
	responses={404: {"description": "Not found"}},
)


@docs_find_and_replace_router.post(
	urls.get("excel").get("find_and_replace"),
	include_in_schema=True,
)
async def find_and_replace(
		background_tasks: BackgroundTasks,
		file: UploadFile = File(...),
		find_value: str = Form(..., max_length=1000),
		replace_value: str = Form(..., max_length=1000),
		token_data: bool = Depends(verify_token),
) -> FileResponse:
	"""
	Attention: This API does not keep the cell formatting. It finds a cell value
	and replaces it with a new value.

	Use `--EMPTY--` to replace a cell value with an empty cell or to find
	empty cells.
	"""
	if find_value == replace_value:
		raise HTTPException(
			status_code=400,
			detail="Find and Replace values cannot be the same."
		)

	validate_doc_file_input(file=file, file_extensions=('xlsx', 'xls'))
	input_file_path = get_temp_file_path(extension='xlsx')
	temp_dir = os.path.dirname(input_file_path)

	async with aiofiles.open(input_file_path, 'wb') as input_file:
		content = await file.read()
		await input_file.write(content)

	try:
		workbook = load_workbook(filename=input_file_path)
	except Exception as e:
		logger.error(f"Could not read the file: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise HTTPException(
			status_code=400,
			detail="Could not read the file."
		)

	try:
		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]

			# Iterate over each cell in the sheet
			for row in sheet.iter_rows():
				for cell in row:
					cell_value = cell.value
					if cell_value is None:
						continue

					if isinstance(cell_value, float) and cell_value.is_integer():
						cell_value = int(cell_value)

					if str(cell_value) == find_value:
						cell.value = replace_value

		output_file_path = os.path.join(temp_dir, f"output.xlsx")
		workbook.save(output_file_path)

		background_tasks.add_task(cleanup_temp_dir, temp_dir=temp_dir)

		return FileResponse(
			output_file_path, filename="output.xlsx", media_type=EXCEL_MEDIA_TYPE
		)
	except HTTPException as e:
		logger.error(f"Error: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise e
	except Exception as e:
		logger.error(f"Error: {e}")
		cleanup_temp_dir(temp_dir=temp_dir)
		raise HTTPException(status_code=500, detail="Server error")
